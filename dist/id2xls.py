#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Standard Library
import os
import sys
import glob
import importlib
import argparse

# Python Package
import openpyxl

# Local Package
from TEST.test import test
from submit import KEY

parser = argparse.ArgumentParser()
parser.add_argument('--id', '-i', default='IDs.txt')
parser.add_argument('--xls', '-x', default='check.xlsx')
parser.add_argument('--dev', action='store_true', default=False)
args = parser.parse_args()


def setCellStr( cell, str ):
    cell.value = str
    cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'code'
wb.create_sheet('check')
wb.create_sheet('count')

ws_code = wb['code']
ws_check = wb['check']
ws_count = wb['count']

wss = [ ws_code, ws_check, ws_count ]

with open(args.id, 'r') as fin:
    IDs = [ line.strip() for line in fin ]
IDs.sort()

for row, ID in enumerate(IDs, start=2):
    for ws in wss:
        setCellStr( ws.cell( row=row, column=1 ), ID )


files = glob.glob('TEST/test?*.py')
files = [ os.path.basename( file ) for file in files ]
files.sort()

for col, file in enumerate(files, start=2):
    print( file )

    for ws in wss:
        setCellStr( ws.cell( row=1, column=col ), file )
    for row, ID in enumerate(IDs, start=2):
        try:
            code, accesskey = test( file, ID, KEY, '', False )
            setCellStr( ws_code.cell( row=row, column=col ), code + ',' + accesskey )
        except:
            pass

wb.save( args.xls )

